import torch
import torch.nn as nn
import torch.nn.functional as F
from torchvision import datasets,transforms
from torch.utils.data import DataLoader
from torch import optim
import matplotlib.pyplot as plt 
from sklearn.metrics import classification_report,accuracy_score,confusion_matrix
import numpy as np 
import os
import seaborn as sns 
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix


data_dir ='C://ML//CNN//chest_xray'
# Transformations: resize, tensor, normalize
transform = transforms.Compose([
                                transforms.Grayscale(),             # Convert to 1 channel (grayscale)
                                transforms.Resize((28,28)),           # Resize to 28x28
                                transforms.ToTensor(),              # Convert to tensor
                                transforms.Normalize((0.5),(0.5))    # Normalize
])

train_data = datasets.ImageFolder(os.path.join(data_dir, "train"),transform=transform)
test_data = datasets.ImageFolder(os.path.join(data_dir, "test"),transform=transform)

#Data Loader
train_loader = DataLoader(train_data,batch_size=64,shuffle=True)
test_loader = DataLoader(test_data,batch_size=64,shuffle=True)

print('Classes:',train_data.classes)


# -------------------------
# 1. Improved CNN Model
# -------------------------
class PneumoniaCNN(nn.Module):
    def __init__(self):
        super(PneumoniaCNN, self).__init__()

        self.conv1 = nn.Sequential(
            nn.Conv2d(1, 32, kernel_size=3, padding=1),
            nn.BatchNorm2d(32),
            nn.ReLU(),
            nn.MaxPool2d(2, 2)
        )
        self.conv2 = nn.Sequential(
            nn.Conv2d(32, 64, kernel_size=3, padding=1),
            nn.BatchNorm2d(64),
            nn.ReLU(),
            nn.MaxPool2d(2, 2)
        )
        self.conv3 = nn.Sequential(
            nn.Conv2d(64, 128, kernel_size=3, padding=1),
            nn.BatchNorm2d(128),
            nn.ReLU(),
            nn.MaxPool2d(2, 2)
        )

        # 🔹 Use dummy input to auto-calc flattened size
        dummy_input = torch.zeros(1, 1, 28, 28)   # (batch, channels, H, W)
        out = self._forward_conv(dummy_input)
        n_features = out.view(1, -1).size(1)

        self.fc1 = nn.Linear(n_features, 256)
        self.dropout = nn.Dropout(0.5)
        self.fc2 = nn.Linear(256, 2)

    def _forward_conv(self, x):
        x = self.conv1(x)
        x = self.conv2(x)
        x = self.conv3(x)
        return x

    def forward(self, x):
        x = self._forward_conv(x)
        x = x.view(x.size(0), -1)   # flatten
        x = F.relu(self.fc1(x))
        x = self.dropout(x)
        x = self.fc2(x)
        return x


# -------------------------
# 2. Training Setup
# -------------------------
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model = PneumoniaCNN().to(device)

criterion = nn.CrossEntropyLoss()

# Adam optimizer
optimizer = optim.Adam(model.parameters(), lr=0.001)

# Learning rate scheduler
scheduler = optim.lr_scheduler.StepLR(optimizer, step_size=5, gamma=0.5)

# -------------------------
# 3. Training Loop (example)
# -------------------------
def train_model(model, train_loader, val_loader, epochs=10):
    for epoch in range(epochs):
        model.train()
        running_loss = 0.0
        correct, total = 0, 0

        for images, labels in train_loader:
            images, labels = images.to(device), labels.to(device)

            optimizer.zero_grad()
            outputs = model(images)

            loss = criterion(outputs, labels)
            loss.backward()
            optimizer.step()

            running_loss += loss.item()

            _, predicted = torch.max(outputs, 1)
            total += labels.size(0)
            correct += (predicted == labels).sum().item()

        scheduler.step()

        train_acc = 100 * correct / total
        print(f"Epoch [{epoch+1}/{epochs}], "
              f"Loss: {running_loss/len(train_loader):.4f}, "
              f"Train Acc: {train_acc:.2f}%")

    return model

trained_model = train_model(model, train_loader, test_loader, epochs=10)

def test_model(model, test_loader):
        model.eval()
        running_loss = 0.0
        y_true, y_pred = [], []

        with torch.no_grad():
            for images, labels in test_loader:
                images, labels = images.to(device), labels.to(device)

                outputs = model(images)
                _, predicted = torch.max(outputs, 1)

                y_true.extend(labels.cpu().numpy())
                y_pred.extend(predicted.cpu().numpy())

       # Compute accuracy
        acc = accuracy_score(y_true, y_pred)
        print(f"\nTest Accuracy: {acc*100:.2f}%")

    # Classification report
        print("\nClassification Report:")
        print(classification_report(y_true, y_pred, target_names=test_loader.dataset.classes))
        # Confusion matrix
        cm = confusion_matrix(y_true, y_pred)
        plt.figure(figsize=(5,4))
        sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",
                xticklabels=test_loader.dataset.classes,
                yticklabels=test_loader.dataset.classes)
        plt.xlabel("Predicted")
        plt.ylabel("True")
        plt.title("Confusion Matrix")
        plt.show()

test_model(model, test_loader) 

#Prediction on single image 
def evaluate_and_export_with_images(model, test_loader, class_names, output_excel="predictions_with_cm.xlsx"):
    model.eval()
    all_preds, all_labels, all_probs, all_paths = [], [], [], []

    with torch.no_grad():
        for images, labels in test_loader:
            images, labels = images.to(device), labels.to(device)
            outputs = model(images)
            probs = F.softmax(outputs, dim=1)

            preds = torch.argmax(probs, dim=1)

            all_preds.extend(preds.cpu().numpy())
            all_labels.extend(labels.cpu().numpy())
            all_probs.extend(probs.cpu().numpy())

            if hasattr(test_loader.dataset, "samples"):
                batch_paths = [p[0] for p in test_loader.dataset.samples[len(all_paths):len(all_paths)+len(labels)]]
                all_paths.extend(batch_paths)

    # ----------------------------
    # 2. Create DataFrame
    # ----------------------------
    df = pd.DataFrame({
        "Image": all_paths,
        "TrueLabel": [class_names[l] for l in all_labels],
        "Predicted": [class_names[p] for p in all_preds],
        "Confidence": [max(prob) for prob in all_probs]
    })

    for i, cls in enumerate(class_names):
        df[f"Prob_{cls}"] = [prob[i] for prob in all_probs]

    # ----------------------------
    # 3. Save to Excel
    # ----------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"

    headers = ["Thumbnail"] + df.columns.tolist()
    ws.append(headers)

    thumb_size = (64, 64)

    for idx, row in df.iterrows():
        excel_row = [None] + row.tolist()
        ws.append(excel_row)

        true_label = row["TrueLabel"]
        pred_label = row["Predicted"]
        pred_cell = ws[f"D{idx+2}"]

        if true_label == pred_label:
            pred_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        else:
            pred_cell.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

        try:
            img = Image.open(row["Image"])
            img.thumbnail(thumb_size)
            thumb_path = f"thumb_{idx}.png"
            img.save(thumb_path)

            xl_img = XLImage(thumb_path)
            xl_img.width, xl_img.height = thumb_size
            ws.add_image(xl_img, f"A{idx+2}")

            os.remove(thumb_path)
        except Exception as e:
            print(f"⚠️ Could not insert image for {row['Image']}: {e}")
# ----------------------------
    # 4. Accuracy + Chart
    # ----------------------------
    total = len(df)
    correct = sum(df["TrueLabel"] == df["Predicted"])
    acc = correct / total

    ws["L1"] = "Accuracy"
    ws["L2"] = acc

    chart = BarChart()
    chart.title = "Class-wise Prediction Count"
    chart.x_axis.title = "Class"
    chart.y_axis.title = "Count"

    counts = df.groupby("Predicted").size().reset_index(name="Count")

    start_row = len(df) + 4
    ws[f"B{start_row}"] = "Class"
    ws[f"C{start_row}"] = "Count"

    for i, row in counts.iterrows():
        ws[f"B{start_row + i + 1}"] = row["Predicted"]
        ws[f"C{start_row + i + 1}"] = row["Count"]

    data_ref = Reference(ws, min_col=3, min_row=start_row, max_row=start_row+len(counts))
    cats_ref = Reference(ws, min_col=2, min_row=start_row+1, max_row=start_row+len(counts))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"L4")

    # ----------------------------
    # 5. Confusion Matrix Heatmap
    # ----------------------------
    cm = confusion_matrix(all_labels, all_preds)
    plt.figure(figsize=(6, 5))
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues", xticklabels=class_names, yticklabels=class_names)
    plt.xlabel("Predicted")
    plt.ylabel("True")
    plt.title("Confusion Matrix")

    cm_path = "confusion_matrix.png"
    plt.savefig(cm_path, bbox_inches="tight")
    plt.close()

    cm_img = XLImage(cm_path)
    cm_img.width, cm_img.height = 400, 300
    ws.add_image(cm_img, f"L20")

    os.remove(cm_path)

    # ----------------------------
    # 6. Save File
    # ----------------------------
    wb.save(output_excel)
    print(f"✅ Predictions + thumbnails + confusion matrix exported to {output_excel}")


evaluate_and_export_with_images(model, test_loader, train_data.classes)

